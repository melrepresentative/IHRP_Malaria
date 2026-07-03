import io
import types
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
from datetime import datetime, date

# Excel-safe helpers
def row_to_excel_safe(row):
    """Convert a dataframe_to_rows row into Excel-safe values."""
    out = []
    for v in row:
        if pd.isna(v):  # catches pd.NA, NaN, None
            out.append(None)
        elif isinstance(v, np.generic):  # numpy scalar -> native Python
            out.append(v.item())
        else:
            out.append(v)
    return out

# NOTE: This Streamlit app should be launched with: streamlit run <this_file.py>
# If someone runs it with plain python, st.stop() may not halt execution, so we also hard-exit.
rules_mod = None
st.set_page_config(page_title="🦟 Malaria Rule-Based Data Validation and Indicator App", layout="wide")
st.title("🦟 Malaria Rule-Based Data Validation and Indicator App")

st.header("Rule file")
rule_file = st.file_uploader("Upload rule file", type=["py"], key="rule_uploader")

if rule_file:
    try:
        code = rule_file.read().decode("utf-8")
        mod = types.ModuleType("user_rules")
        exec(compile(code, "user_rules.py", "exec"), mod.__dict__)
        rules_mod = mod

        available_rule_names = [
            name for name in dir(mod)
            if callable(getattr(mod, name)) and name.startswith("malaria_")
        ]


        if not available_rule_names:
            st.error("❌ No malaria_* rule functions found in the uploaded rule file.")
            st.stop()

        st.success(f"✅ Rules loaded: {', '.join(sorted(available_rule_names))}")

    except Exception as e:
        st.error("❌ Failed to load rule file.")
        st.exception(e)
        st.stop()

else:
    st.warning("⚠️ Please upload a rule file (.py) to continue.")
    st.stop()

def to_excel_safe(df: pd.DataFrame) -> pd.DataFrame:
    """Convert df to types that openpyxl can write without errors."""
    out = df.copy()
    out = out.where(pd.notna(out), None)
    for c in out.columns:
        dt = out[c].dtype
        dt_name = str(dt)
        if dt_name.startswith(("Int", "UInt", "boolean")) or pd.api.types.is_sparse(dt):
            out[c] = out[c].astype(object)

        # ✅ SCREENING_DATE export fix:
        # - Keep invalid strings unchanged
        # - Convert Timestamp/datetime to python date (no time part) so Excel won't show 00:00:00

        if str(c).strip().casefold() == "screening_date":
            def _as_date(v):
                if v is None or pd.isna(v):
                    return None
                if isinstance(v, pd.Timestamp):
                    return v.date()
                if isinstance(v, datetime):
                    return v.date()
                return v  # keep text exactly as entered (including invalid dates)

            out[c] = out[c].map(_as_date)
    return out

def put_comment_first(df: pd.DataFrame) -> pd.DataFrame:
    return df if "COMMENT" not in df.columns else df[["COMMENT"] + [c for c in df.columns if c != "COMMENT"]]

def _df_display_without_time(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for col, dtype in out.dtypes.items():
        if pd.api.types.is_datetime64_any_dtype(dtype):
            out[col] = pd.to_datetime(out[col], errors="coerce").dt.strftime("%d-%b-%y")
    return out

def _build_error_summary_counts(comment_series: pd.Series) -> pd.DataFrame:
    """
    Error Type Summary
    Rules:
      - Split COMMENT by ';'
      - Group normal messages into standard buckets
      - Count duplicate/retesting pairs only once without pair keys:
          * "Duplicate with row X"
          * "Retesting with row X"
        are counted under "Duplicate/Retesting"
    """
    counts = {}

    for comment in comment_series.astype(str):
        parts = [p.strip() for p in comment.split(";") if p.strip()]

        for p in parts:
            msg = p.strip()
            msg_lower = msg.lower()

            # DUPLICATE / RETESTING
            # Count one pair once, without pair key column
            if "duplicate with row" in msg_lower or "retesting with row" in msg_lower:
                key = "Duplicate/Retesting"
                counts[key] = counts.get(key, 0) + 0.5
                continue

            # NORMAL ERROR TYPES
            if msg_lower.endswith(" missing"):
                key = "Missing"

            elif msg_lower.endswith(" invalid format"):
                key = "Date invalid format"

            elif msg_lower.endswith(" out of range"):
                key = "Out of range"

            elif msg_lower.endswith(" not integer"):
                key = "Not integer"

            elif msg_lower.endswith(" invalid option"):
                key = "Invalid option"

            elif msg_lower.endswith(" invalid"):
                key = "Invalid"

            elif " vs " in msg_lower and "inconsistent" in msg_lower:
                key = "Inconsistent"

            else:
                key = "Other"

            counts[key] = counts.get(key, 0) + 1

    if not counts:
        return pd.DataFrame(columns=["Error Type", "Count"])

    # convert 0.5 + 0.5 pair totals into integers
    rows = []
    rows = [(k, int(v)) for k, v in counts.items()]

    return (
        pd.DataFrame(rows, columns=["Error Type", "Count"])
        .sort_values("Count", ascending=False)
        .reset_index(drop=True)
    )

def _build_error_detail_counts(comment_series: pd.Series) -> pd.DataFrame:
    """
    Counts full error messages (e.g., 'Consistency_Check Error: [X vs Y]').
    """
    counts = {}
    for comment in comment_series.astype(str):
        parts = [p.strip() for p in comment.split(";") if p.strip()]
        for p in parts:
            counts[p] = counts.get(p, 0) + 1

    if not counts:
        return pd.DataFrame(columns=["Error Type Detail", "Count"])

    return (
        pd.DataFrame(list(counts.items()), columns=["Error Type Detail", "Count"])
        .sort_values("Count", ascending=False)
        .reset_index(drop=True)
    )

def _build_remark_reach_summary(remark_series: pd.Series) -> pd.DataFrame:
    """
    Counts Remark (REACH) values, including blanks as '(blank)'.
    """
    s = remark_series.copy() if remark_series is not None else pd.Series(dtype=object)
    s = s.astype(object)

    def norm(v):
        if pd.isna(v) or str(v).strip() == "":
            return "(blank)"
        return str(v).strip()

    s2 = s.map(norm)
    vc = s2.value_counts(dropna=False)

    return vc.rename_axis("Remark (REACH) Summary").reset_index(name="Count")

st.markdown("---")

# === Single combined flow: Cleaning + Indicators via rule file ===
st.header("Upload and Process Workbook")
excel_file = st.file_uploader("Upload your malaria Excel file", type=["xlsx", "xls"], key="workbook_uploader")

if excel_file is not None:
    try:
        # Cache original bytes BEFORE reading with pandas
        original_bytes = excel_file.getvalue() if hasattr(excel_file, "getvalue") else excel_file.read()
        xls = pd.ExcelFile(io.BytesIO(original_bytes))
        sheet_names = xls.sheet_names
    except Exception as e:
        st.error("❌ Could not read the Excel file.")
        st.exception(e)
        xls = None
        sheet_names = []

    if sheet_names:
    # ✅ Keep one-sheet logic
    # ✅ Show both "Process multiple sheets" and "ALL sheets" directly
        if len(sheet_names) == 1:
            selected = sheet_names[:]  # auto-select the only sheet
            st.info(f"Sheet to process: {sheet_names[0]}")
        else:
            multi = st.checkbox("Process multiple sheets", value=False, key="proc_multi")
            select_all = st.checkbox("ALL sheets", value=False, key="proc_select_all")

            if select_all:
                selected = sheet_names[:]
                st.info(f"All sheets selected: {', '.join(sheet_names)}")
            elif multi:
                selected = st.multiselect(
                    "Select sheet(s) to process",
                    sheet_names,
                    key="proc_select_ms"
                )
            else:
                one = st.selectbox(
                    "Select sheet to process",
                    sheet_names,
                    key="proc_select_sb"
                )
                selected = [one]  # keep list so your for-loop still works

        st.markdown("<br><br>", unsafe_allow_html=True)
        
        # --- Rule selection per sheet (works even when sheet names are random) ---
        available_rule_names = [
            name for name in dir(rules_mod)
            if callable(getattr(rules_mod, name)) and name.startswith("malaria_")
        ]

        sheet_rule_choice = {}

        if selected:
            if available_rule_names:
                st.markdown("### Select rule(s) for each sheet")
                default_rule = "malaria_positive" if "malaria_positive" in available_rule_names else available_rule_names[0]

                for sh in selected:
                    sheet_rule_choice[sh] = st.multiselect(
                        f"Rule(s) for sheet: {sh}",
                        options=available_rule_names,
                        default=[default_rule],
                        key=f"rule_for_{sh}"
                    )
            else:
                st.warning("No malaria_* rule functions found in the rule file.")

        run_btn = st.button("Run", key="run_all")

        if run_btn:
            if not selected:
                st.warning("Select at least one sheet, then click **Run**.")
            else:
                processed = {}
                sheet_summaries = {}
                errors = []

                for sheet in selected:
                    try:
                        raw_df = xls.parse(sheet_name=sheet)
                        chosen_rule_names = sheet_rule_choice.get(sheet, [])

                        if not chosen_rule_names:
                            raise Exception(f"No rule selected for sheet: {sheet}")

                        for chosen_rule_name in chosen_rule_names:
                            try:
                                # malaria_recheck is a workbook-level rule
                                if chosen_rule_name == "malaria_recheck":
                                    continue

                                rule_func = getattr(rules_mod, chosen_rule_name, None)

                                if not callable(rule_func):
                                    raise Exception(f"Rule function not found: {chosen_rule_name}")

                                out_df = rule_func(raw_df.copy())

                                processed_key = f"{sheet} | {chosen_rule_name}"

                                processed[processed_key] = {
                                    "display": put_comment_first(out_df.copy()),
                                    "file": out_df.copy(),
                                    "rule_name": chosen_rule_name,
                                    "source_sheet": sheet,
                                }

                                # Safer COMMENT extraction (keeps index aligned)
                                comments = out_df.get("COMMENT", pd.Series("", index=out_df.index))

                                sheet_summaries[processed_key] = {
                                    "type_summary": _build_error_summary_counts(comments),
                                    "detail_summary": _build_error_detail_counts(comments),
                                    "remark_summary": _build_remark_reach_summary(
                                        out_df["Remark (REACH)"] if "Remark (REACH)" in out_df.columns else pd.Series(dtype=object)
                                    ),
                                }

                            except Exception as e:
                                errors.append((f"{sheet} | {chosen_rule_name}", e))

                    except Exception as e:
                        errors.append((sheet, e))

                        
                # Build REACH_RECHECK only if user selected malaria_recheck for any chosen sheet ---
                wants_recheck = any(
                    "malaria_recheck" in sheet_rule_choice.get(s, [])
                    for s in selected
                )

                if wants_recheck and hasattr(rules_mod, "malaria_recheck"):
                    original_input_columns = set()
                    positive_df_for_recheck = None
                    aggregate_df_for_recheck = None

                    # 1) First try to use already-processed outputs
                    for sname, bundle in processed.items():
                        if bundle.get("rule_name") == "malaria_positive" and positive_df_for_recheck is None:
                            positive_df_for_recheck = bundle["file"].copy()
                            original_input_columns = set(bundle["file"].columns)

                        if bundle.get("rule_name") == "malaria_aggregate" and aggregate_df_for_recheck is None:
                            aggregate_df_for_recheck = bundle["file"].copy()

                    # 2) If still missing, try to build from selected ORIGINAL sheets
                    def looks_like_positive(df: pd.DataFrame) -> bool:
                        cols = {str(c).strip().upper() for c in df.columns}

                        has_patient = "PATIENT_NAME" in cols
                        has_date = "SCREENING_DATE" in cols
                        has_result = ("RDT" in cols) or ("MICROSCOPY" in cols) or ("RDT_MICROSCOPY" in cols)

                        # CMHN normally has FACILITY_NAME
                        # MFM normally has STATE_REGION + TOWNSHIP
                        has_cmhn_key = "FACILITY_NAME" in cols
                        has_mfm_key = "STATE_REGION" in cols and "TOWNSHIP" in cols

                        return has_patient and has_date and has_result and (has_cmhn_key or has_mfm_key)

                    def looks_like_aggregate(df: pd.DataFrame) -> bool:
                        cols = {str(c).strip().upper() for c in df.columns}
                        required = {"AGE_<5", "AGE_5-14", "AGE>=15"}
                        return required.issubset(cols)

                    if positive_df_for_recheck is None or aggregate_df_for_recheck is None:
                        for s in selected:
                            raw_df = xls.parse(sheet_name=s)

                            if positive_df_for_recheck is not None and aggregate_df_for_recheck is not None:
                                break

                            if positive_df_for_recheck is None and looks_like_positive(raw_df):
                                original_input_columns = set(raw_df.columns)

                                # Prefer original data if it already has CMHN/MFM indicator columns.
                                # Otherwise process it with malaria_positive so RECHECK has CM1_REACH etc.
                                raw_cols = {str(c).strip().upper() for c in raw_df.columns}
                                has_indicator = any(
                                    c in raw_cols
                                    for c in {
                                        "CM1_REACH",
                                        "CM1_CMHN",
                                        "CM1_MFM",
                                        "CM2_D_REACH",
                                        "CM2_D_CMHN",
                                        "CM2_D_MFM",
                                        "CM2_N_REACH",
                                        "CM2_N_CMHN",
                                        "CM2_N_MFM",
                                    }
                                )

                                if has_indicator:
                                    positive_df_for_recheck = raw_df.copy()
                                else:
                                    positive_df_for_recheck = rules_mod.malaria_positive(raw_df.copy())

                                # RECHECK needs AGE_GP for pivot.
                                if "AGE_GP" not in positive_df_for_recheck.columns and "AGE_YEAR" in positive_df_for_recheck.columns:
                                    age_num = pd.to_numeric(positive_df_for_recheck["AGE_YEAR"], errors="coerce")
                                    positive_df_for_recheck["AGE_GP"] = np.select(
                                        [
                                            age_num.lt(5),
                                            age_num.between(5, 14, inclusive="both"),
                                            age_num.ge(15),
                                        ],
                                        ["<5", "5-14", ">=15"],
                                        default=""
                                    )

                            if aggregate_df_for_recheck is None and looks_like_aggregate(raw_df):
                                aggregate_df_for_recheck = rules_mod.malaria_aggregate(raw_df.copy())

                    # 3) Build REACH_RECHECK
                    if positive_df_for_recheck is not None and aggregate_df_for_recheck is not None:
                        try:
                            processed["REACH_RECHECK"] = rules_mod.malaria_recheck(
                                positive_df_for_recheck.copy(),
                                aggregate_df_for_recheck.copy(),
                                original_input_columns,
                            )
                        except Exception as e:
                            errors.append(("REACH_RECHECK", e))
                    else:
                        errors.append((
                            "REACH_RECHECK",
                            Exception(
                                "REACH_RECHECK requires one positive source sheet and one aggregate source sheet among the selected sheets."
                            )
                        ))
                             
                # Preview
                if processed:
                    st.subheader("👀 Preview (first 10 rows)")
                    tabs = st.tabs(list(processed.keys()))
                    for tab, sheet in zip(tabs, processed.keys()):
                        with tab:
                            st.dataframe(
                                _df_display_without_time(processed[sheet]["display"].head(10)),
                                use_container_width=True,
                            )

                    # Error summary combined table
                    st.subheader("📊 Error Summary (by sheet)")
                    for sheet, bundle in sheet_summaries.items():
                        st.markdown(f"**{sheet}**")

                        df_type = bundle["type_summary"]
                        df_detail = bundle["detail_summary"]
                        df_remark = bundle["remark_summary"]

                        st.markdown("**Error Type Summary**")
                        st.dataframe(df_type, use_container_width=True)

                        st.markdown("**Error Type Detail**")
                        st.dataframe(df_detail, use_container_width=True)


                if errors:
                    st.warning("Some sheets failed to process:")
                    for sheet, e in errors:
                        with st.expander(f"Details: {sheet}"):
                            st.exception(e)

                # Build output workbook: ADD processed sheets (keep originals intact) + add one "Error Summary" sheet
                if processed:
                    st.subheader("⬇️ Download")
                    try:
                        from openpyxl import Workbook
                        wb = Workbook()
                        orig_wb = load_workbook(io.BytesIO(original_bytes))

                        # Remove the default empty sheet
                        default_ws = wb.active
                        wb.remove(default_ws)

                        # Add "Processed - <sheet>" sheets without touching original ones
                        for sheet_name, bundle in processed.items():
                            if "sheet_title" in bundle:
                                new_title = bundle["sheet_title"]
                            else:
                                new_title = f"Processed - {sheet_name}"

                            # Ensure uniqueness
                            title = new_title
                            suffix = 1
                            while title in wb.sheetnames:
                                suffix += 1
                                title = f"{new_title} ({suffix})"

                            ws = wb.create_sheet(title=title)  # append at end
                            if "custom_rows" in bundle:
                                for r in bundle["custom_rows"]:
                                    ws.append(row_to_excel_safe(r))
                            else:
                                safe_df = to_excel_safe(bundle["file"])
                                for r in dataframe_to_rows(safe_df, index=False, header=True):
                                    ws.append(row_to_excel_safe(r))
                            # Copy SCREENING_DATE formatting only for sheets that exist in original workbook
                            source_sheet_name = bundle.get("source_sheet", sheet_name)

                            if source_sheet_name in orig_wb.sheetnames:
                                orig_ws = orig_wb[source_sheet_name]

                                # Find SCREENING_DATE column in BOTH sheets using header row (row 1)
                                def find_col_index(sheet, target_name: str):
                                    target = target_name.strip().casefold()
                                    for col in range(1, sheet.max_column + 1):
                                        v = sheet.cell(row=1, column=col).value
                                        if v is None:
                                            continue
                                        if str(v).strip().casefold() == target:
                                            return col
                                    return None

                                proc_date_col = find_col_index(ws, "SCREENING_DATE")
                                orig_date_col = find_col_index(orig_ws, "SCREENING_DATE")

                                if proc_date_col and orig_date_col:
                                    max_rows = min(ws.max_row, orig_ws.max_row)

                                    for rr in range(2, max_rows + 1):
                                        src_cell = orig_ws.cell(row=rr, column=orig_date_col)
                                        dst_cell = ws.cell(row=rr, column=proc_date_col)

                                        dst_cell.value = src_cell.value
                                        dst_cell.number_format = src_cell.number_format

                        # --- Table 1: Error Type Summary (with Sheet column) ---
                        rows_type = []
                        for sname, bundle in sheet_summaries.items():
                            df_sum = bundle["type_summary"]
                            if not df_sum.empty:
                                for _, rr in df_sum.iterrows():
                                    rows_type.append([sname, rr["Error Type"], int(rr["Count"])])

                        df_type_summary = pd.DataFrame(rows_type, columns=["Sheet", "Error Type", "Count"])

                        # --- Table 2: Error Type Detail (with Sheet column) ---
                        rows_detail = []
                        for sname, bundle in sheet_summaries.items():
                            df_detail = bundle["detail_summary"]
                            if not df_detail.empty:
                                for _, rr in df_detail.iterrows():
                                    rows_detail.append([sname, rr["Error Type Detail"], int(rr["Count"])])

                        df_detail_all = (
                            pd.DataFrame(rows_detail, columns=["Sheet", "Error Type Detail", "Count"])
                            .sort_values(["Sheet", "Count"], ascending=[True, False])
                            .reset_index(drop=True)
                        )

                        # --- Table 3: Remark (REACH) Summary (combined across sheets) ---
                        remark_counts = {}
                        for sname, bundle in sheet_summaries.items():
                            df_rem = bundle["remark_summary"]
                            for _, rr in df_rem.iterrows():
                                k = rr["Remark (REACH) Summary"]
                                remark_counts[k] = remark_counts.get(k, 0) + int(rr["Count"])

                        df_remark_all = (
                            pd.DataFrame(list(remark_counts.items()), columns=["Remark (REACH) Summary", "Count"])
                            .sort_values("Count", ascending=False)
                            .reset_index(drop=True)
                        )

                        # Place "Error Summary" as the FIRST sheet if possible
                        # If a sheet with that name exists, replace its contents; otherwise create it.
                        if "Error Summary" in wb.sheetnames:
                            ws_sum = wb["Error Summary"]
                            ws_sum.delete_rows(1, ws_sum.max_rows)
                        else:
                            ws_sum = wb.create_sheet(title="Error Summary", index=0)

                        def write_df(ws, df, title, start_row):
                            ws.cell(row=start_row, column=1, value=title)
                            start_row += 1

                            # headers
                            for j, col in enumerate(df.columns, start=1):
                                ws.cell(row=start_row, column=j, value=col)
                            start_row += 1

                            # rows
                            for row in df.itertuples(index=False):
                                for j, v in enumerate(row, start=1):
                                    ws.cell(row=start_row, column=j, value=v)
                                start_row += 1

                            return start_row

                        r = 1
                        r = write_df(ws_sum, df_type_summary, "Error Type Summary", r) + 2
                        r = write_df(ws_sum, df_detail_all, "Error Type Detail", r) + 2
                        r = write_df(ws_sum, df_remark_all, "Remark (REACH) Summary", r) + 2

                        out = io.BytesIO()
                        wb.save(out)
                        out.seek(0)

                        st.download_button(
                            label="📥 Download (Processed sheets + Error Summary)",
                            data=out.getvalue(),
                            file_name="malaria_processed_with_summary.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )
                    except Exception as e:
                        st.error("❌ Failed to generate the processed workbook.")
                        st.exception(e)

# NOTE:
# - Only Proceed sheets will get. 
# - 'Error Summary' is added/updated as the first sheet.

